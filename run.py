import speech_recognition as sr
import openpyxl
import os
import customtkinter
import time
import pyttsx3
import threading
import re
import json
import pyaudio
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from threading import Thread
from playsound3 import playsound as play
from PIL import Image, ImageTk, ImageDraw
from vosk import Model, KaldiRecognizer

def load_config(file_path):
    try:
        with open(file_path, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {"variables": {}, "grupos": [], "image_path": "default_image.png"}

def save_config(file_path, data):
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

config_file = "config.json"
config_data = load_config(config_file)

def update_variable(key, value):
    config_data["variables"][key] = value #Dentro del grupo variables, en el parametro x, agrega
    save_config(config_file, config_data)

#Sacar valores del array grupos en el config.json
#grupos = config_data["grupos"]

grupo_actual = str(config_data["variables"].get("grupo", ""))
fecha_actual = str(config_data["variables"].get("fecha", ""))

#Clear console
def clear():
    os.system('cls' if os.name == 'nt' else 'clear')

#Crear un reconocedor
r = sr.Recognizer()
model = Model("model-es-small")
recognizer = KaldiRecognizer(model, 16000)

i = 0
alumnos = []

#Cargar lista de alumnos y elementos de db
conn = sqlite3.connect('students.db')
cursor = conn.cursor()
lista = cursor.execute(f"SELECT * FROM students WHERE grupo='{grupo_actual}'").fetchall()
no_alumnos = cursor.execute(f"SELECT COUNT(*) FROM students WHERE grupo='{grupo_actual}'").fetchone()[0]
fechas = cursor.execute("SELECT DISTINCT fecha FROM students").fetchall()
grupos = cursor.execute("SELECT DISTINCT grupo FROM students").fetchall()
#conn.close()

#Convertir cada tupla en una lista para poder modificarla
lista = [list(tupla) for tupla in lista]

#Cargar archivo de alumnos
try:
    with open("alumnos_raw.txt", "r") as file:
        existing_strings = set(file.read().splitlines())  #Cargar lineas en un set
        no_alumnos = len(existing_strings)  #Cuenta el numero de lineas
except FileNotFoundError:
    existing_strings = set()  #Si no existe el archivo, crear un nuevo set
alumnos_sort = sorted(existing_strings)

exit_event = threading.Event()
exit_event_retardos = threading.Event()

def thread_start(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label, info_timer_label):
    t = Thread(target=pase_lista, args=(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label, info_timer_label,), daemon=True)
    t.start()

def thread_retardos(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label):
    t = Thread(target=retardos, args=(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label,), daemon=True)
    t.start()

def thread_countdown(info_timer_label):
    t = Thread(target=countdown, args=(info_timer_label), daemon=True)
    t.start()

def thread_hora_actual():
    t = Thread(target=tiempo_actual, daemon=True)
    t.start()

def tiempo_actual():
    hora_actual_label = None
    while True:
        hora = datetime.now().strftime("%H:%M:%S")
        if hora_actual_label:
            hora_actual_label.destroy()
        hora_actual_label = customtkinter.CTkLabel(master=main_frame_top_bar, text=hora, text_color="#75003E", font=("Roboto Regular", 16, "bold"))
        hora_actual_label.pack(padx=10, pady=3, side="right", expand=True)
        time.sleep(1)

def pase_lista(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label, info_timer_label):
    grupo_actual = str(config_data["variables"].get("grupo", ""))

    #Números del 1 al 25 en texto y su valor numérico
    numeros_texto = {
        "uno": 1, "dos": 2, "tres": 3, "cuatro": 4, "cinco": 5, "seis": 6, "siete": 7, "ocho": 8, "nueve": 9,
        "diez": 10, "once": 11, "doce": 12, "trece": 13, "catorce": 14, "quince": 15, "dieciseis": 16, "diecisiete": 17,
        "dieciocho": 18, "diecinueve": 19, "veinte": 20, "veintiuno": 21, "veintidos": 22, "veintidós": 22, "veintitres": 23, "veintitrés": 23, "veinticuatro": 24, "veinticinco": 25
    }

    for i in range(no_alumnos):
        lista[i][5] = "Falta"
        print(lista[i][5])
        i=+1

    button_continuar.configure(command=lambda:[thread_retardos(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label), thread_countdown(info_timer_label), exit_event.set()])

    #Comienza pase de lista

    #Start audio stream
    mic = pyaudio.PyAudio()
    stream = mic.open(format=pyaudio.paInt16, channels=1, rate=16000, input=True, frames_per_buffer=8000)
    stream.start_stream()
    print("Escuchando...")
    
    info_listening_label.configure(text="Escuchando...", text_color="#45CE30")
    
    while True:
        data = stream.read(4000, exception_on_overflow=False)
        
        if recognizer.AcceptWaveform(data):
            result = recognizer.FinalResult()
            palabras = str(result)
            cleaned_str = re.sub('["{}:"]|text', '', palabras)
            cleaned_str = cleaned_str.split()
            print("Oración Reconocida: ", cleaned_str)

            #Comprobar si el número aparece como palabra o como número
            for i, numero in enumerate(numeros_texto):
                if numero in cleaned_str or str(numeros_texto[numero]) in cleaned_str:
                    print(f"Variable Número: {numero}")
                    print(f"Variable cleaned_str: {cleaned_str}")
                    numero_reconocido = int(numeros_texto[numero])
                    print(f"Otra Variable: {numero_reconocido}")
                    info_listening_label.configure(text=(f"Registrando Asistencia Para #{numero_reconocido}"), text_color="#F3B63A")
                    numero_reconocido = numero_reconocido - 1
                    lista[numero_reconocido][5] = "Puntual"
                    play("./assets/sounds/puntual.mp3")                    

            info_listening_label.configure(text="Escuchando...", text_color="#45CE30")

            #Terminar Thread
            if exit_event.is_set():
                #Imprimir los resultados
                for i in range(no_alumnos):
                    alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][0]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                    alumno_label.grid(row=i, column=0, sticky="W", padx=10)
                    alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][2]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                    alumno_label.grid(row=i, column=1, sticky="W", padx=20)
                    alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][1]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                    alumno_label.grid(row=i, column=2, sticky="W", padx=20)
                    alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][3]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                    alumno_label.grid(row=i, column=3, sticky="W", padx=20)
                    alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][4]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                    alumno_label.grid(row=i, column=4, sticky="W", padx=20)
                    if lista[i][5] == "Puntual":
                        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][5]}"), text_color="#45CE30", font=("Roboto Regular", 16, "bold"))
                        alumno_label.grid(row=i, column=5, sticky="W", padx=20)
                    #print(f"Número: {lista[i][0]} | Apellido: {lista[i][2]} | Nombre: {lista[i][1]} |  Grupo: {lista[i][3]} | Boleta: {lista[i][4]} | Estado: {lista[i][5]}")
                    i =+ 1
                print("Thread Pase Lista Terminado")
                break
        else:
            result = recognizer.PartialResult()
            continue

def retardos(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label):
    info_listening_label.configure(text="Comienzan a Contar Retardos", text_color="#F3B63A")
    time.sleep(3)
    #Inicia tiempo de tolerancia para retardos

    #Números del 1 al 25 en texto y su valor numérico
    numeros_texto = {
        "uno": 1, "dos": 2, "tres": 3, "cuatro": 4, "cinco": 5, "seis": 6, "siete": 7, "ocho": 8, "nueve": 9,
        "diez": 10, "once": 11, "doce": 12, "trece": 13, "catorce": 14, "quince": 15, "dieciseis": 16, "diecisiete": 17,
        "dieciocho": 18, "diecinueve": 19, "veinte": 20, "veintiuno": 21, "veintidos": 22, "veintidós": 22, "veintitres": 23, "veintitrés": 23, "veinticuatro": 24, "veinticinco": 25
    }

    button_continuar.configure(command=lambda:[guardar_lista_db(), exit_event_retardos.set()], text="Terminar")

    #Start audio stream
    mic = pyaudio.PyAudio()
    stream = mic.open(format=pyaudio.paInt16, channels=1, rate=16000, input=True, frames_per_buffer=8000)
    stream.start_stream()
    info_listening_label.configure(text="Escuchando...", text_color="#45CE30")
    print("Escuchando...")

    while True:
        data = stream.read(4000, exception_on_overflow=False)
        #textbox.delete("0.0", "end")
        
        if recognizer.AcceptWaveform(data):
            result = recognizer.FinalResult()
            palabras = str(result)
            cleaned_str = re.sub('["{}:"]|text', '', palabras)
            #cleaned_str = ' '.join(cleaned_str.split())
            cleaned_str = cleaned_str.split()
            print("Oración Reconocida: ", cleaned_str)

            #Comprobar si el número aparece como palabra o como número
            for i, numero in enumerate(numeros_texto):
                if numero in cleaned_str or str(numeros_texto[numero]) in cleaned_str:
                    print(f"Variable Número: {numero}")
                    print(f"Variable cleaned_str: {cleaned_str}")
                    numero_reconocido = int(numeros_texto[numero])
                    print(f"Otra Variable: {numero_reconocido}")
                    info_listening_label.configure(text=(f"Registrando Retardo Para #{numero_reconocido}"), text_color="#F3B63A")
                    numero_reconocido = numero_reconocido - 1
                    lista[numero_reconocido][5] = "Retardo"
                    play("./assets/sounds/puntual.mp3")
            
                    info_listening_label.configure(text="Escuchando...", text_color="#45CE30")
                    if exit_event_retardos.is_set(): #Si se presiona el botón que termina el thread retardos
                        #Imprimir los resultados
                        for i in range(no_alumnos):
                            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][0]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                            alumno_label.grid(row=i, column=0, sticky="W", padx=10)
                            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][2]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                            alumno_label.grid(row=i, column=1, sticky="W", padx=20)
                            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][1]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                            alumno_label.grid(row=i, column=2, sticky="W", padx=20)
                            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][3]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                            alumno_label.grid(row=i, column=3, sticky="W", padx=20)
                            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][4]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
                            alumno_label.grid(row=i, column=4, sticky="W", padx=20)
                            if lista[i][5] == "Retardo":
                                alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][5]}"), text_color="#F3B63A", font=("Roboto Regular", 16, "bold"))
                                alumno_label.grid(row=i, column=5, sticky="W", padx=20)
                            #print(f"Número: {lista[i][0]} | Apellido: {lista[i][2]} | Nombre: {lista[i][1]} |  Grupo: {lista[i][3]} | Boleta: {lista[i][4]} | Estado: {lista[i][5]}")
                            i =+ 1
                        print("Thread Retardos Terminado")
                        break
                else:
                    result = recognizer.PartialResult()
                    continue

def countdown(info_timer_label):
    tiempo_tolerancia = int(config_data["variables"]["tiempo_tolerancia"]) * 60 #Carga el parametro tiempo_tolerancia y lo convierte a int
    time.sleep(5)
    info_timer_label = None
    while tiempo_tolerancia:
        mins, secs = divmod(tiempo_tolerancia, 60)
        timer_tolerancia = mins
        timer = '{:02d}:{:02d}'.format(mins, secs)
        if timer_tolerancia <= (1):
            time_running_out = "red"
        else:
            time_running_out = "#45CE30"
        if info_timer_label:  # Si ya existe, destruir el label previo antes de crear uno nuevo
            info_timer_label.destroy()
        info_timer_label.configure(text=timer, text_color=time_running_out)
        time.sleep(1) 
        tiempo_tolerancia -= 1
        if tiempo_tolerancia < 1:
            exit_event_retardos.set()
            if exit_event_retardos.is_set():
                print("Thread Countdown Terminado")
                break
    if info_timer_label:  # Si ya existe, destruir el label previo antes de crear uno nuevo
            info_timer_label.destroy()
    time.sleep(5)
    guardar_lista_db()

def guardar_lista_db():
    exit_event_retardos.set()
    current_date = datetime.now().strftime("%d/%m/%Y")
    #Cargar lista de alumnos y elementos de db
    conn = sqlite3.connect('students.db')
    cursor = conn.cursor()
    alumnos = []
    for i in range(no_alumnos):
        alumnos.append([lista[i][0], lista[i][1], lista[i][2], lista[i][3], lista[i][4], lista[i][5], current_date])
        i =+ 1
    cursor.executemany("INSERT INTO students values (?,?,?,?,?,?,?)", alumnos)
    conn.commit()
    print(alumnos)

def guardar_lista_excel():
    #exit_event_retardos.set()
    time.sleep(1)
    
    current_date = datetime.now().strftime("%d/%m/%Y %H:%M")

    custom_workbook = openpyxl.load_workbook("./assets/Lista de Asistencia.xlsx")
    sheet = custom_workbook.active

    sheet['B4'] = "ESCUELA SUPERIOR DE INGENIERÍA MECÁNICA Y ELECTRÍCA UNIDAD CULHUACÁN"
    sheet['B5'] = "NOMBRE DEL DOCENTE: CABRERA TEJEDA JUAN JOSÉ"
    sheet['C6'] = current_date

    dia_actual = datetime.now().day
        
    columnas = list(range(4, 34))  #D es la columna 4 y AH es la columna 34

    #Imprimir los nombres
    fila_inicio = 8
    for i, alumno in enumerate(no_alumnos):
        if fila_inicio + i <= 42:  # Para no exceder la fila 42
            sheet.cell(row=fila_inicio + i, column=3).value = alumno

    # Verificar si la celda de la fila 7 contiene el día actual
    for col in columnas:
        celda = sheet.cell(row=7, column=col)
        if celda.value == dia_actual:
            # Si se encuentra la columna con el día actual, escribir la información de numeros[]
            for i, asistencia in enumerate(no_alumnos):
                if fila_inicio + i <= 42:  # Para no exceder la fila 42
                    if asistencia == 'Puntual':
                        shortener = 'P'
                        sheet.cell(row=fila_inicio + i, column=col).value = shortener
                        sheet.cell(row=fila_inicio + i, column=col).fill = PatternFill(start_color='019031', end_color='019031', fill_type='solid')
                        sheet.cell(row=fila_inicio + i, column=col).font = Font(color='FFFFFF')
                        print(f"Escribiendo en fila {fila_inicio + i}, columna {col}: {shortener}")
                    elif asistencia == 'Retardo':
                        shortener = 'R'
                        sheet.cell(row=fila_inicio + i, column=col).value = shortener
                        sheet.cell(row=fila_inicio + i, column=col).fill = PatternFill(start_color='F3B431', end_color='F3B431', fill_type='solid')
                        sheet.cell(row=fila_inicio + i, column=col).font = Font(color='FFFFFF')
                        print(f"Escribiendo en fila {fila_inicio + i}, columna {col}: {shortener}")
                    else:
                        shortener = 'F'
                        sheet.cell(row=fila_inicio + i, column=col).value = shortener
                        sheet.cell(row=fila_inicio + i, column=col).fill = PatternFill(start_color='B83227', end_color='B83227', fill_type='solid')
                        sheet.cell(row=fila_inicio + i, column=col).font = Font(color='FFFFFF')
                        print(f"Escribiendo en fila {fila_inicio + i}, columna {col}: {shortener}")
            break  # Salir del bucle una vez que se escribe en la columna correcta
    current_date = datetime.now().strftime("%d-%m-%Y %H-%M")
    custom_workbook.save(f"Lista {current_date}.xlsx")

def gestionar_alumno(no, nombre, apellido, grupo, boleta):

    #Buscar Alumno
    alumno = cursor.execute(f"""SELECT * FROM students WHERE
                            (number='{no}') OR
                            (first_name='{nombre}') OR
                            (last_name='{apellido}') OR
                            (grupo='{grupo}') OR
                            (boleta='{boleta}')""").fetchall()
    
    no_asistencias = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND ((estado='Puntual') OR (estado='Retardo'))").fetchone()[0]
    no_faltas = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND (estado='Falta')").fetchone()[0]
    no_asistencias_puntual = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND (estado='Puntual')").fetchone()[0]
    no_asistencias_retardo = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND (estado='Retardo')").fetchone()[0]
    registros_totales = no_asistencias + no_faltas
    porcentaje_asistencias = int((no_asistencias/(registros_totales))*100)


    print(f"Asistencias: {no_asistencias}\nPuntuales: {no_asistencias_puntual}\nRetardos: {no_asistencias_retardo}\nFaltas: {no_faltas}\nPorcentaje: {porcentaje_asistencias}%")
    if no_faltas >= 3:
        print("Problema")
    else:
        print("En orden")

    delete_pages()

    #Top Frame
    config_page_frame_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_frame_top_bar.pack_propagate(False)
    config_page_frame_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    config_page_frame_top_bar.configure(height=60)

    #Top Frame Label
    config_page_frame_top_label = customtkinter.CTkLabel(master=config_page_frame_top_bar, text=(f"{alumno[0][1]} {alumno[0][2]}"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    config_page_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    config_page_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#F3F4F7")
    config_page_frame.pack(padx=0, pady=0, expand=True, side="top", fill="both")

    #Data Frame
    info_frame = customtkinter.CTkFrame(master=config_page_frame, fg_color="#FFFFFF")
    info_frame.pack_propagate(True)
    info_frame.pack(padx=0, pady=0, fill="both")
    info_frame.configure(height=60)
    info_label = customtkinter.CTkLabel(master=info_frame, text=("Asistencias"), text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=("Puntuales"), text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=("Retardos"), text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=("Faltas"), text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=("Porcentaje de Asistencias"), text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=("Status"), text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)

    info_frame = customtkinter.CTkFrame(master=config_page_frame, fg_color="#FFFFFF")
    info_frame.pack_propagate(True)
    info_frame.pack(padx=0, pady=0, fill="both")
    info_frame.configure(height=60)
    info_label = customtkinter.CTkLabel(master=info_frame, text=no_asistencias, text_color="black", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=no_asistencias_puntual, text_color="#45CE30", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=no_asistencias_retardo, text_color="#F3B63A", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=no_faltas, text_color="red", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text=(f"{porcentaje_asistencias}%"), text_color="black", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="N/A", text_color="black", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    if no_faltas >= 3:
        info_label.configure(text="Problema", text_color="red")
    else:
        info_label.configure(text="En Orden", text_color="#45CE30")

    #Alumno Frame Individual
    alumno_frame = customtkinter.CTkScrollableFrame(master=config_page_frame, fg_color="#FFFFFF")
    alumno_frame.pack_propagate(True)
    alumno_frame.pack(padx=0, pady=0, fill="both")
    alumno_frame.configure(height=600)

    

    for i in range(registros_totales):
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][0]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=0, sticky="W", padx=10)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][2]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=1, sticky="W", padx=10)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][1]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=2, sticky="W", padx=20)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][3]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=3, sticky="W", padx=20)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][4]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=4, sticky="W", padx=20)
        if lista[i][5] == "Puntual":
            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][5]}"), text_color="#45CE30", font=("Roboto Regular", 16, "bold"))
            alumno_label.grid(row=i, column=5, sticky="W", padx=20)
        elif lista[i][5] == "Retardo":
            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][5]}"), text_color="#F3B63A", font=("Roboto Regular", 16, "bold"))
            alumno_label.grid(row=i, column=5, sticky="W", padx=20)
        else:
            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][5]}"), text_color="red", font=("Roboto Regular", 16, "bold"))
            alumno_label.grid(row=i, column=5, sticky="W", padx=20)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{alumno[i][6]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=6, sticky="W", padx=20)
        i =+ 1

def editar_alumno(no, nombre, apellido, grupo, boleta):

    #Buscar Alumno
    alumno = cursor.execute(f"""SELECT * FROM students WHERE
                            (number='{no}') OR
                            (first_name='{nombre}') OR
                            (last_name='{apellido}') OR
                            (grupo='{grupo}') OR
                            (boleta='{boleta}')""").fetchall()
    
    no_asistencias = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND ((estado='Puntual') OR (estado='Retardo'))").fetchone()[0]
    no_faltas = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND (estado='Falta')").fetchone()[0]
    no_asistencias_puntual = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND (estado='Puntual')").fetchone()[0]
    no_asistencias_retardo = cursor.execute(f"SELECT COUNT (*) FROM students WHERE (last_name='{alumno[0][2]}') AND (estado='Retardo')").fetchone()[0]
    porcentaje_asistencias = int((no_asistencias/(no_asistencias + no_faltas))*100)

    print(f"Asistencias: {no_asistencias}\nPuntuales: {no_asistencias_puntual}\nRetardos: {no_asistencias_retardo}\nFaltas: {no_faltas}\nPorcentaje: {porcentaje_asistencias}%")
    if no_faltas >= 3:
        print("Problema")
    else:
        print("En orden")

    delete_pages()

    #Top Frame
    config_page_frame_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_frame_top_bar.pack_propagate(False)
    config_page_frame_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    config_page_frame_top_bar.configure(height=60)

    #Top Frame Label
    config_page_frame_top_label = customtkinter.CTkLabel(master=config_page_frame_top_bar, text=(f"{alumno[0][1]} {alumno[0][2]}"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    config_page_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    config_page_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#F3F4F7")
    config_page_frame.pack(padx=0, pady=0, expand=True, side="top", fill="both")

    
    config_page_button_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_button_frame.pack(padx=0, pady=0, expand=False, side="bottom", fill="x")
    config_page_button_frame.configure(height=120)

    #Numero
    no_label = customtkinter.CTkLabel(master=config_page_frame, text="No", text_color="black", font=("Roboto Regular", 16, "bold"))
    no_label.pack()
    no_entry = customtkinter.CTkEntry(master=config_page_frame)
    no_entry.insert(0, alumno[0][0])
    no_entry.pack()

    #Nombre
    nombre_label = customtkinter.CTkLabel(master=config_page_frame, text="Nombre", text_color="black", font=("Roboto Regular", 16, "bold"))
    nombre_label.pack()
    nombre_entry = customtkinter.CTkEntry(master=config_page_frame)
    nombre_entry.insert(0, alumno[0][1])
    nombre_entry.pack()

    #Apellido
    apellido_label = customtkinter.CTkLabel(master=config_page_frame, text="Apellido", text_color="black", font=("Roboto Regular", 16, "bold"))
    apellido_label.pack()
    apellido_entry = customtkinter.CTkEntry(master=config_page_frame)
    apellido_entry.insert(0, alumno[0][2])
    apellido_entry.pack()

    #Grupo
    grupo_label = customtkinter.CTkLabel(master=config_page_frame, text="Grupo", text_color="black", font=("Roboto Regular", 16, "bold"))
    grupo_label.pack()
    grupo_entry = customtkinter.CTkEntry(master=config_page_frame)
    grupo_entry.insert(0, alumno[0][3])
    grupo_entry.pack()

    #Boleta
    boleta_label = customtkinter.CTkLabel(master=config_page_frame, text="Boleta", text_color="black", font=("Roboto Regular", 16, "bold"))
    boleta_label.pack()
    boleta_entry = customtkinter.CTkEntry(master=config_page_frame)
    boleta_entry.insert(0, alumno[0][4])
    boleta_entry.pack()

    button_guardar = customtkinter.CTkButton(master=config_page_button_frame,
                                                fg_color="#75003E",
                                                compound="left",
                                                text="Guardar",
                                                text_color="#FFFFFF",
                                                font=("Roboto Regular", 16, "bold"),
                                                corner_radius=5,
                                                width=160,
                                                height=40,
                                                command=lambda:
                                                [   
                                                    gestionar_alumno(no_entry.get(),
                                                                     nombre_entry.get(),
                                                                     apellido_entry.get(),
                                                                     grupo_entry.get(),
                                                                     boleta_entry.get())
                                                ])
    button_guardar.place(relx=0.38, rely=0.3)

def frase_del_dia():
    #frase_del_dia_main_label.tag_config('puntual', foreground="#45CE30")
    
    #Start audio stream
    mic = pyaudio.PyAudio()
    stream = mic.open(format=pyaudio.paInt16, channels=1, rate=16000, input=True, frames_per_buffer=8000)
    stream.start_stream()
    print("Escuchando...")

    while True:
        data = stream.read(4000, exception_on_overflow=False)
        #textbox.delete("0.0", "end")
        
        if recognizer.AcceptWaveform(data):
            result = recognizer.FinalResult()
            palabras = str(result)
            cleaned_str = re.sub('["{}:"]|text', '', palabras)
            #cleaned_str = ' '.join(cleaned_str.split())
            cleaned_str = cleaned_str.split()
            print("Oración Reconocida: ", cleaned_str)
        else:
            result = recognizer.PartialResult()
            continue

def salir_programa():
    quit()
     
app = customtkinter.CTk()
app.title("Speech Recognition App")
app.geometry("1280x768")

def delete_pages():
    for frame in main_frame.winfo_children():
        frame.destroy()

def ver_lista_page():
    delete_pages()

    grupo_actual = config_data["variables"]["grupo"]
    fecha_actual = config_data["variables"]["fecha"]
    lista = cursor.execute(f"SELECT * FROM students WHERE fecha='{fecha_actual}'").fetchall()
    no_alumnos = cursor.execute(f"SELECT COUNT(*) FROM students WHERE fecha='{fecha_actual}'").fetchone()[0]

    #Top Frame
    ver_lista_frame_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    ver_lista_frame_top_bar.pack_propagate(False)
    ver_lista_frame_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    ver_lista_frame_top_bar.configure(height=60)

    #Top Frame Label
    ver_lista_frame_top_label = customtkinter.CTkLabel(master=ver_lista_frame_top_bar, text=(f"Grupo: {grupo_actual} - Fecha: {fecha_actual}"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    ver_lista_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    ver_lista_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#FFFFFF")
    ver_lista_frame.pack_propagate(True)
    ver_lista_frame.pack(padx=0, pady=0, fill="both")

    #Alumno Frame
    info_frame = customtkinter.CTkFrame(master=ver_lista_frame, fg_color="#FFFFFF")
    info_frame.pack_propagate(True)
    info_frame.pack(padx=0, pady=0, fill="both")
    info_frame.configure(height=60)
    info_label = customtkinter.CTkLabel(master=info_frame, text="No", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=False)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Apellido", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Nombre", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Grupo", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Boleta", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Estado", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Fecha", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)

    alumno_frame = customtkinter.CTkScrollableFrame(master=ver_lista_frame, fg_color="#FFFFFF")
    alumno_frame.pack_propagate(True)
    alumno_frame.pack(padx=0, pady=0, fill="both")
    alumno_frame.configure(height=600)


    for i in range(no_alumnos):
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][0]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=0, sticky="W", padx=10)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][2]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=1, sticky="W", padx=10)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][1]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=2, sticky="W", padx=20)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][3]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=3, sticky="W", padx=20)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][4]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=4, sticky="W", padx=20)
        if lista[i][5] == "Puntual":
            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][5]}"), text_color="#45CE30", font=("Roboto Regular", 16, "bold"))
            alumno_label.grid(row=i, column=5, sticky="W", padx=20)
        elif lista[i][5] == "Retardo":
            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][5]}"), text_color="#F3B63A", font=("Roboto Regular", 16, "bold"))
            alumno_label.grid(row=i, column=5, sticky="W", padx=20)
        else:
            alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][5]}"), text_color="red", font=("Roboto Regular", 16, "bold"))
            alumno_label.grid(row=i, column=5, sticky="W", padx=20)
        alumno_label = customtkinter.CTkLabel(master=alumno_frame, text=(f"{lista[i][6]}"), text_color="black", font=("Roboto Regular", 16, "bold"))
        alumno_label.grid(row=i, column=6, sticky="W", padx=20)
        i =+ 1
    
def pasar_lista_page():
    delete_pages()

    #Top Frame
    pasar_lista_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    pasar_lista_top_bar.pack_propagate(False)
    pasar_lista_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    pasar_lista_top_bar.configure(height=60)

    #Top Frame Label
    pasar_lista_frame_top_label = customtkinter.CTkLabel(master=pasar_lista_top_bar, text=(f"Pasar Asistencia"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    pasar_lista_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    pasar_lista_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#B6B6B6")
    pasar_lista_frame.pack_propagate(False)
    pasar_lista_frame.pack(padx=0, pady=0, side="top", fill="both", expand=True)
    pasar_lista_frame.configure(height=600)

    #Data Frame
    info_frame = customtkinter.CTkFrame(master=pasar_lista_frame, fg_color="#FFFFFF")
    info_frame.pack_propagate(False)
    info_frame.pack(padx=0, pady=0, fill="both")
    info_frame.configure(height=60)
    info_listening_label = customtkinter.CTkLabel(master=info_frame, text="Escuchando: Deshabilitado", text_color="#B6B6B6", font=("Roboto Regular", 20, "bold"))
    info_listening_label.pack(padx=10, pady=3, side="left", expand=True)
    info_timer_label = customtkinter.CTkLabel(master=info_frame, text="Timer: Deshabilitado", text_color="#B6B6B6", font=("Roboto Regular", 20, "bold"))
    info_timer_label.pack(padx=10, pady=3, side="left", expand=True)
    
    info_frame = customtkinter.CTkFrame(master=pasar_lista_frame, fg_color="#FFFFFF")
    info_frame.pack_propagate(True)
    info_frame.pack(padx=0, pady=0, fill="both")
    info_frame.configure(height=60)
    info_label = customtkinter.CTkLabel(master=info_frame, text="No", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=False)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Apellido", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Nombre", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Grupo", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Boleta", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)
    info_label = customtkinter.CTkLabel(master=info_frame, text="Estado", text_color="#B6B6B6", font=("Roboto Regular", 16, "bold"))
    info_label.pack(padx=10, pady=3, side="left", expand=True)

    #Lista de alumnos frame
    alumno_frame = customtkinter.CTkScrollableFrame(master=pasar_lista_frame, fg_color="#FFFFFF")
    alumno_frame.pack_propagate(True)
    alumno_frame.configure(height=600)
    alumno_frame.pack(padx=0, pady=0, side="top", fill="both", expand=True)

    textbox = customtkinter.CTkTextbox(master=pasar_lista_frame, fg_color="transparent", text_color="#404040")
    textbox.configure(font=('Roboto', 20), height=600)
    textbox.pack(expand=True, side="right", fill="both")

    detener_paso_lista_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    detener_paso_lista_frame.pack_propagate(False)
    detener_paso_lista_frame.pack(padx=0, pady=0, expand=False, side="bottom", fill="x")
    detener_paso_lista_frame.configure(height=120)

    button_continuar = customtkinter.CTkButton(master=detener_paso_lista_frame,
                                            fg_color="#75003E",
                                            compound="left",
                                            text="Comenzar Retardos",
                                            text_color="#FFFFFF",
                                            font=("Roboto Regular", 16, "bold"),
                                            corner_radius=5,
                                            width=160,
                                            height=40)
    button_continuar.place(relx=0.38, rely=0.3)
    thread_start(textbox, button_continuar, pasar_lista_frame, info_label, alumno_frame, info_listening_label, info_timer_label)

def frase_del_dia_page():
    delete_pages()
    frase_del_dia = config_data["variables"]["frase_del_dia"]

    #Top Frame
    frase_del_dia_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    frase_del_dia_top_bar.pack_propagate(False)
    frase_del_dia_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    frase_del_dia_top_bar.configure(height=60)

    #Top Frame Label
    frase_del_dia_frame_top_label = customtkinter.CTkLabel(master=frase_del_dia_top_bar, text=("Frase del Dia"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    frase_del_dia_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    frase_del_dia_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#FFFFFF")
    frase_del_dia_frame.pack_propagate(False)
    frase_del_dia_frame.pack(padx=0, pady=0, fill="both")

    #Frase del dia
    frase_del_dia_main_label = customtkinter.CTkLabel(master=frase_del_dia_frame, text=(frase_del_dia), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    frase_del_dia_main_label.pack(padx=10, pady=3, side="left", expand=True)

    frase_del_dia()

def gestionar_page():
    delete_pages()

    #Top Frame
    config_page_frame_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_frame_top_bar.pack_propagate(False)
    config_page_frame_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    config_page_frame_top_bar.configure(height=60)

    #Top Frame Label
    config_page_frame_top_label = customtkinter.CTkLabel(master=config_page_frame_top_bar, text=(f"Gestionar Estudiantes"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    config_page_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    config_page_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#F3F4F7")
    config_page_frame.pack(padx=0, pady=0, expand=True, side="top", fill="both")

    
    config_page_button_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_button_frame.pack(padx=0, pady=0, expand=False, side="bottom", fill="x")
    config_page_button_frame.configure(height=120)

    #Entries

    #Numero
    no_label = customtkinter.CTkLabel(master=config_page_frame, text="No", text_color="black", font=("Roboto Regular", 16, "bold"))
    no_label.pack()
    no_entry = customtkinter.CTkEntry(master=config_page_frame)
    no_entry.insert(0, "")
    no_entry.pack()

    #Nombre
    nombre_label = customtkinter.CTkLabel(master=config_page_frame, text="Nombre", text_color="black", font=("Roboto Regular", 16, "bold"))
    nombre_label.pack()
    nombre_entry = customtkinter.CTkEntry(master=config_page_frame)
    nombre_entry.insert(0, "")
    nombre_entry.pack()

    #Apellido
    apellido_label = customtkinter.CTkLabel(master=config_page_frame, text="Apellido", text_color="black", font=("Roboto Regular", 16, "bold"))
    apellido_label.pack()
    apellido_entry = customtkinter.CTkEntry(master=config_page_frame)
    apellido_entry.insert(0, "")
    apellido_entry.pack()

    #Grupo
    grupo_label = customtkinter.CTkLabel(master=config_page_frame, text="Grupo", text_color="black", font=("Roboto Regular", 16, "bold"))
    grupo_label.pack()
    grupo_entry = customtkinter.CTkEntry(master=config_page_frame)
    grupo_entry.insert(0, "")
    grupo_entry.pack()

    #Boleta
    boleta_label = customtkinter.CTkLabel(master=config_page_frame, text="Boleta", text_color="black", font=("Roboto Regular", 16, "bold"))
    boleta_label.pack()
    boleta_entry = customtkinter.CTkEntry(master=config_page_frame)
    boleta_entry.insert(0, "")
    boleta_entry.pack()

    button_analizar = customtkinter.CTkButton(master=config_page_button_frame,
                                                fg_color="#75003E",
                                                compound="left",
                                                text="Buscar",
                                                text_color="#FFFFFF",
                                                font=("Roboto Regular", 16, "bold"),
                                                corner_radius=5,
                                                width=160,
                                                height=40,
                                                command=lambda:
                                                [   
                                                    gestionar_alumno(no_entry.get(),
                                                                     nombre_entry.get(),
                                                                     apellido_entry.get(),
                                                                     grupo_entry.get(),
                                                                     boleta_entry.get())
                                                ])
    button_analizar.pack()
    button_editar = customtkinter.CTkButton(master=config_page_button_frame,
                                                fg_color="#75003E",
                                                compound="left",
                                                text="Editar",
                                                text_color="#FFFFFF",
                                                font=("Roboto Regular", 16, "bold"),
                                                corner_radius=5,
                                                width=160,
                                                height=40,
                                                command=lambda:
                                                [   
                                                    editar_alumno(no_entry.get(),
                                                                     nombre_entry.get(),
                                                                     apellido_entry.get(),
                                                                     grupo_entry.get(),
                                                                     boleta_entry.get())
                                                ])
    button_editar.pack()

def config_page():
    delete_pages()

    #Top Frame
    config_page_frame_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_frame_top_bar.pack_propagate(False)
    config_page_frame_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
    config_page_frame_top_bar.configure(height=60)

    #Top Frame Label
    config_page_frame_top_label = customtkinter.CTkLabel(master=config_page_frame_top_bar, text=(f"Configuración"), text_color="#75003E", font=("Roboto Regular", 20, "bold"))
    config_page_frame_top_label.pack(padx=10, pady=3, side="left", expand=False)

    #Main Frame
    config_page_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#F3F4F7")
    config_page_frame.pack(padx=0, pady=0, expand=True, side="top", fill="both")

    
    config_page_button_frame = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
    config_page_button_frame.pack(padx=0, pady=0, expand=False, side="bottom", fill="x")
    config_page_button_frame.configure(height=120)

    # Cambiar variables

    ###Nombre
    #nombre_label = customtkinter.CTkLabel(master=config_page_frame, text="Nombre de Bienvenida", text_color="black", font=("Roboto Regular", 16, "bold"))
    #ombre_label.pack()
    #nombre_entry = customtkinter.CTkEntry(master=config_page_frame)
    #nombre_entry.insert(0, str(config_data["variables"].get("nombre", "")))
    #nombre_entry.pack()

    #Tolerancia
    tolerancia_label = customtkinter.CTkLabel(master=config_page_frame, text="Tolerancia (mins)", text_color="black", font=("Roboto Regular", 16, "bold"))
    tolerancia_label.pack()
    tolerancia_entry = customtkinter.CTkEntry(master=config_page_frame)
    tolerancia_entry.insert(0, str(config_data["variables"].get("tiempo_tolerancia", "")))
    tolerancia_entry.pack()

    #Seleccionar Grupos
    grupos_label = customtkinter.CTkLabel(master=config_page_frame, text="Seleccionar Grupo", text_color="black", font=("Roboto Regular", 16, "bold"))
    grupos_label.pack()
    
    grupos_dropdown = customtkinter.CTkComboBox(master=config_page_frame, values=[grupo_actual])
    grupos_valores = [grupo[0] for grupo in grupos]  # Extraer solo los valores
    grupos_dropdown.configure(values=grupos_valores)
    grupos_dropdown.pack()

    #Seleccionar Fecha
    fechas_label = customtkinter.CTkLabel(master=config_page_frame, text="Seleccionar Fecha", text_color="black", font=("Roboto Regular", 16, "bold"))
    fechas_label.pack()
    
    fechas_dropdown = customtkinter.CTkComboBox(master=config_page_frame, values=[fecha_actual])
    fechas_valores = [fecha[0] for fecha in fechas]  # Extraer solo los valores
    fechas_dropdown.configure(values=fechas_valores)
    fechas_dropdown.pack()

    #Modelo de Reconocimiento
    modelo_label = customtkinter.CTkLabel(master=config_page_frame, text="Seleccionar Modelo", text_color="black", font=("Roboto Regular", 16, "bold"))
    modelo_label.pack()
    modelo_dropdown = customtkinter.CTkComboBox(master=config_page_frame,
                                                values=["Español (Light)",
                                                        "Español (Full)",
                                                        "English (Light)",
                                                        "English (Full)"])
    modelo_dropdown.pack()

    #Frase del dia
    frase_del_dia_label = customtkinter.CTkLabel(master=config_page_frame, text="Frase del Día", text_color="black", font=("Roboto Regular", 16, "bold"))
    frase_del_dia_label.pack()
    frase_del_dia_entry = customtkinter.CTkEntry(master=config_page_frame)
    frase_del_dia_entry.insert(0, str(config_data["variables"].get("frase_del_dia", "")))
    frase_del_dia_entry.pack()

    button_actualizar = customtkinter.CTkButton(master=config_page_button_frame,
                                                fg_color="#75003E",
                                                compound="left",
                                                text="Guardar Cambios",
                                                text_color="#FFFFFF",
                                                font=("Roboto Regular", 16, "bold"),
                                                corner_radius=5,
                                                width=160,
                                                height=40,
                                                command=lambda:
                                                [   
                                                    update_variable("tiempo_tolerancia", tolerancia_entry.get()),
                                                    update_variable("frase_del_dia", frase_del_dia_entry.get()),
                                                    update_variable("grupo", grupos_dropdown.get()),
                                                    update_variable("fecha", fechas_dropdown.get()),
                                                 
                                                ])
    button_actualizar.place(relx=0.38, rely=0.3)

#Dia, hora y tiempo del dia
dia_interfaz = datetime.now().strftime("%d/%m/%Y")
hora_interfaz = datetime.now().strftime("%H:%M:%S")
hora_actual = current_date = datetime.now().hour

if hora_actual >= 20:
    hora_actual = "Buenas noches, profesor"
elif hora_actual >= 12:
    hora_actual = "Buenas tardes, profesor"
else:
    hora_actual = "Buenos días, profesor"


ver_lista_icon = customtkinter.CTkImage(light_image=Image.open("./assets/icons/ver_lista.png"), size=(30, 30))
pasar_lista_icon = customtkinter.CTkImage(light_image=Image.open("./assets/icons/pasar_lista.png"), size=(30, 30))
salir_icon = customtkinter.CTkImage(light_image=Image.open("./assets/icons/salir.png"), size=(30, 30))

logo_image = Image.open("./assets/images/logo.png")
logo_image = logo_image.resize((80, 120))
logo_image_tk = ImageTk.PhotoImage(logo_image)

logo_esime = Image.open("./assets/images/esime.png")
logo_esime = logo_esime.resize((50, 50))
logo_esime_tk = ImageTk.PhotoImage(logo_esime)

main_image = Image.open("./assets/images/avatar.png")
main_image = main_image.resize((300, 300))  #Tamaño de la imagen
main_image_tk = ImageTk.PhotoImage(main_image)

#main_bg = Image.open("./assets/images/bg_main.png")
#main_bg = main_bg.resize((300, 300))  #Tamaño de la imagen
#main_bg_tk = ImageTk.PhotoImage(main_bg)

sidebar = customtkinter.CTkFrame(master=app, fg_color="#75003E")
sidebar.pack_propagate(False)
sidebar.pack(padx=0, pady=0, side="left", fill="y", expand=False)
sidebar.configure(width=200)

logo_label = customtkinter.CTkLabel(master=sidebar, image=logo_image_tk, text="")
logo_label.pack(pady=(10, 20))

main_frame = customtkinter.CTkFrame(master=app, fg_color="#FFFFFF")
main_frame.pack(padx=0, pady=0, expand=True, side="right", fill="both")

main_frame_top_bar = customtkinter.CTkFrame(master=main_frame, fg_color="#FFE3F3")
main_frame_top_bar.pack_propagate(False)
main_frame_top_bar.pack(padx=0, pady=0, expand=False, side="top", fill="x")
main_frame_top_bar.configure(height=60)

#Dia actual
hora_label = customtkinter.CTkLabel(master=main_frame_top_bar, text=(dia_interfaz), text_color="#75003E", font=("Roboto Regular", 16, "bold"))
hora_label.pack(padx=10, pady=3, side="left", expand=True)

#Hora actual
thread_hora_actual()

#Logo Esime
logo_esime_label = customtkinter.CTkLabel(master=main_frame_top_bar, image=logo_esime_tk, text="")
logo_esime_label.pack(padx=10, pady=3, side="left", expand=True)

#Main image centered
main_image_label = customtkinter.CTkLabel(master=main_frame, image=main_image_tk, text="")
main_image_label.place(relx=0.5, rely=0.5, anchor="center")

#Saludo
main_label = customtkinter.CTkLabel(master=main_frame, text=(hora_actual), text_color="#75003E", font=("Roboto Regular", 40, "bold"))
main_label.pack(pady=20)

#Nombre
profesor_label = customtkinter.CTkLabel(master=main_frame, text="Cabrera Tejeda Juan José", text_color="#75003E", font=("Roboto Regular", 32))
profesor_label.place(relx=0.5, rely=0.83, anchor="center")

buttons_frame = customtkinter.CTkFrame(master=sidebar, fg_color="transparent")
buttons_frame.pack(expand=True, side="top", fill="y")

button = customtkinter.CTkButton(master=buttons_frame,
                                 image=ver_lista_icon,
                                 compound="left",
                                 text="Ver Lista",
                                 text_color="#404040",
                                 font=("Roboto Regular", 16, "bold"),
                                 corner_radius=5,
                                 fg_color="#FFFFFF",
                                 command=ver_lista_page,
                                 width=160,
                                 height=40,
                                 anchor="w")
button.pack(padx=20, pady=20)
button.bind("<Enter>", lambda event: button.configure(text_color="#FFFFFF",
                                                        border_color="#FFFFFF",
                                                        border_width=1,
                                                        fg_color="transparent"))
button.bind("<Leave>", lambda event: button.configure(text_color="#404040",
                                                        fg_color="white"))

button_pasar_lista = customtkinter.CTkButton(master=buttons_frame,
                                  image=pasar_lista_icon,
                                  compound="left",
                                  text="Pasar Lista",
                                  text_color="#404040",
                                  font=("Roboto Regular", 16, "bold"),
                                  corner_radius=5,
                                  fg_color="#FFFFFF",
                                  width=160,
                                  height=40,
                                  anchor="w",
                                  command=pasar_lista_page)
button_pasar_lista.pack(padx=20, pady=20)
button_pasar_lista.bind("<Enter>", lambda event: button_pasar_lista.configure(text_color="#FFFFFF",
                                                        border_color="#FFFFFF",
                                                        border_width=1,
                                                        fg_color="transparent"))
button_pasar_lista.bind("<Leave>", lambda event: button_pasar_lista.configure(text_color="#404040",
                                                        fg_color="white"))

"""button_fdd = customtkinter.CTkButton(master=buttons_frame,
                                  image=pasar_lista_icon,
                                  compound="left",
                                  text="Frase del Dia",
                                  text_color="#404040",
                                  font=("Roboto Regular", 16, "bold"),
                                  corner_radius=5,
                                  fg_color="#FFFFFF",
                                  width=160,
                                  height=40,
                                  anchor="w",
                                  command=frase_del_dia_page)
button_fdd.pack(padx=20, pady=20)
button_fdd.bind("<Enter>", lambda event: button_fdd.configure(text_color="#FFFFFF",
                                                        border_color="#FFFFFF",
                                                        border_width=1,
                                                        fg_color="transparent"))
button_fdd.bind("<Leave>", lambda event: button_fdd.configure(text_color="#404040",
                                                        fg_color="white"))"""

button_gestionar = customtkinter.CTkButton(master=buttons_frame,
                                  image=salir_icon,
                                  compound="left",
                                  text="Gestionar",
                                  text_color="#404040",
                                  font=("Roboto Regular", 16, "bold"),
                                  corner_radius=5,
                                  fg_color="#FFFFFF",
                                  width=160,
                                  height=40,
                                  anchor="w",
                                  command=gestionar_page)
button_gestionar.pack(padx=20, pady=20)
button_gestionar.bind("<Enter>", lambda event: button_gestionar.configure(text_color="#FFFFFF",
                                                        border_color="#FFFFFF",
                                                        border_width=1,
                                                        fg_color="transparent"))
button_gestionar.bind("<Leave>", lambda event: button_gestionar.configure(text_color="#404040",
                                                        fg_color="white"))

button_confguracion = customtkinter.CTkButton(master=buttons_frame,
                                  image=salir_icon,
                                  compound="left",
                                  text="Configuración",
                                  text_color="#404040",
                                  font=("Roboto Regular", 16, "bold"),
                                  corner_radius=5,
                                  fg_color="#FFFFFF",
                                  width=160,
                                  height=40,
                                  anchor="w",
                                  command=config_page)
button_confguracion.pack(padx=20, pady=20)
button_confguracion.bind("<Enter>", lambda event: button_confguracion.configure(text_color="#FFFFFF",
                                                        border_color="#FFFFFF",
                                                        border_width=1,
                                                        fg_color="transparent"))
button_confguracion.bind("<Leave>", lambda event: button_confguracion.configure(text_color="#404040",
                                                        fg_color="white"))

button_salir = customtkinter.CTkButton(master=buttons_frame,
                                  image=salir_icon,
                                  compound="left",
                                  text="Salir",
                                  text_color="#404040",
                                  font=("Roboto Regular", 16, "bold"),
                                  corner_radius=5,
                                  fg_color="#FFFFFF",
                                  width=160,
                                  height=40,
                                  anchor="w",
                                  command=salir_programa)
button_salir.pack(padx=20, pady=20)
button_salir.bind("<Enter>", lambda event: button_salir.configure(text_color="#FFFFFF",
                                                        border_color="#FFFFFF",
                                                        border_width=1,
                                                        fg_color="transparent"))
button_salir.bind("<Leave>", lambda event: button_salir.configure(text_color="#404040",
                                                        fg_color="white"))

app.mainloop()

